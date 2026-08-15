import calendar
import io
import os
import re
import zipfile

from collections import defaultdict
from copy import copy
from datetime import date
from datetime import datetime

from flask import Blueprint
from flask import jsonify
from flask import request
from flask import send_file
from flask import session
from openpyxl import load_workbook

import app as app_pkg

from .auth import elevated_required
from .auth import login_required
from .constants import ELEVATED_ROLES
from .helpers import parse_time

export_bp = Blueprint('exports', __name__)

TEMPLATE_PATH = os.path.join(os.path.dirname(os.path.dirname(__file__)), 'templates', 'Monthly_Worklog_Template.xlsx')
MONTH_NAMES = [
    'January', 'February', 'March', 'April', 'May', 'June',
    'July', 'August', 'September', 'October', 'November', 'December',
]


def _safe_export_name(name):
    value = (name or '').strip()
    safe = re.sub(r'[^\w\-.]', '_', value)
    return safe or 'member'


def _copy_row_style(worksheet, source_row, target_row, num_cols=10):
    """Copy cell styles when the template needs more rows."""
    for column in range(1, num_cols + 1):
        src = worksheet.cell(row=source_row, column=column)
        dst = worksheet.cell(row=target_row, column=column)
        if src.has_style:
            dst.font = copy(src.font)
            dst.border = copy(src.border)
            dst.fill = copy(src.fill)
            dst.alignment = copy(src.alignment)


def generate_excel_bytes(employee_id, member, year, months=None):
    """Generate a filled Excel workbook for one employee and return raw bytes.

    `employee_id` is dbo.Employee.EmployeeID (post-migration); param name is
    kept generic to avoid changing callers.
    """
    first_day = date(year, 1, 1)
    last_day = date(year, 12, 31)
    worklogs = app_pkg.db.query(
        """
        SELECT log_date, project, task, projectdepartment, description,
               start_time, end_time, hours, status, note
        FROM worklogs
        WHERE EmployeeID = ? AND log_date BETWEEN ? AND ?
        ORDER BY log_date, start_time
        """,
        (employee_id, first_day, last_day),
    )

    by_month_day = defaultdict(list)
    for worklog in worklogs:
        # Format here (not in SQL) so the same query string works on both
        # engines — see helpers.parse_time(), which expects a 'HH:MM' str.
        worklog['start_time'] = worklog['start_time'].strftime('%H:%M') if worklog['start_time'] else None
        worklog['end_time'] = worklog['end_time'].strftime('%H:%M') if worklog['end_time'] else None
        log_date = worklog['log_date']
        if isinstance(log_date, str):
            log_date = datetime.strptime(log_date, '%Y-%m-%d').date()
        by_month_day[(log_date.month, log_date.day)].append(worklog)

    workbook = load_workbook(TEMPLATE_PATH)

    dashboard = workbook['Dashboard']
    dashboard['A1'] = f'{year} ANNUAL WORKLOG DASHBOARD'
    dashboard['B2'] = member['name']
    dashboard['B3'] = member['department']
    dashboard['E2'] = member.get('staff_id') or ''
    dashboard['E3'] = member.get('position') or ''

    empty_sheet = workbook['empty']
    data_start = 5
    template_rows = 997

    for sheet_name in MONTH_NAMES:
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    months_to_generate = months if months is not None else list(range(1, 13))
    for month_idx in months_to_generate:
        sheet_name = MONTH_NAMES[month_idx - 1]
        worksheet = workbook.copy_worksheet(empty_sheet)
        worksheet.title = sheet_name
        worksheet['A1'] = f'{sheet_name} {year} - Work Log'

        days_in_month = calendar.monthrange(year, month_idx)[1]
        row_data = []
        for day in range(1, days_in_month + 1):
            current_date = date(year, month_idx, day)
            entries = by_month_day.get((month_idx, day), [])
            for index, entry in enumerate(entries):
                row_data.append((current_date if index == 0 else None, entry))

        needed_rows = len(row_data)
        if needed_rows > template_rows:
            extra_rows = needed_rows - template_rows
            worksheet.insert_rows(data_start + template_rows, extra_rows)
            for row_number in range(data_start + template_rows, data_start + needed_rows):
                _copy_row_style(worksheet, data_start + template_rows - 1, row_number)

        for index, (entry_date, entry) in enumerate(row_data):
            row_number = data_start + index
            if entry_date is not None:
                worksheet.cell(row=row_number, column=1).value = entry_date
                worksheet.cell(row=row_number, column=1).number_format = 'dd/mm/yyyy'
            if entry is None:
                continue

            worksheet.cell(row=row_number, column=2).value = entry['project'] or ''
            worksheet.cell(row=row_number, column=3).value = entry['projectdepartment'] or ''
            worksheet.cell(row=row_number, column=4).value = entry['description'] or ''
            worksheet.cell(row=row_number, column=5).value = entry['task'] or ''
            worksheet.cell(row=row_number, column=6).value = parse_time(entry['start_time'])
            worksheet.cell(row=row_number, column=7).value = parse_time(entry['end_time'])
            if worksheet.cell(row=row_number, column=6).value:
                worksheet.cell(row=row_number, column=6).number_format = 'h:mm'
            if worksheet.cell(row=row_number, column=7).value:
                worksheet.cell(row=row_number, column=7).number_format = 'h:mm'
            if entry.get('hours') is not None:
                worksheet.cell(row=row_number, column=8).value = float(entry['hours'])
            worksheet.cell(row=row_number, column=9).value = entry['status'] or ''
            worksheet.cell(row=row_number, column=10).value = entry['note'] or ''

    if 'empty' in workbook.sheetnames:
        del workbook['empty']

    project_names = [project['name'] for project in app_pkg.db.query("SELECT name FROM projects ORDER BY name")]
    for sheet_name in ('tbl_project', 'Config'):
        if sheet_name in workbook.sheetnames:
            worksheet = workbook[sheet_name]
            for row_number in range(2, worksheet.max_row + 1):
                worksheet.cell(row=row_number, column=1).value = None
            for index, project_name in enumerate(project_names):
                worksheet.cell(row=index + 2, column=1).value = project_name

    buffer = io.BytesIO()
    workbook.save(buffer)
    return buffer.getvalue()


@export_bp.route('/api/export/excel', methods=['GET'])
@login_required
def export_excel():
    member_id = (request.args.get('member_id') or '').strip()
    year = request.args.get('year', type=int, default=date.today().year)
    months_str = request.args.get('months', '')
    months = [int(month) for month in months_str.split(',') if month.strip().isdigit()] or None

    if not member_id:
        return jsonify({'error': 'member_id required'}), 400

    if not app_pkg._worklog_open and session.get('role') not in ELEVATED_ROLES and member_id != session.get('member_id'):
        return jsonify({'error': 'Permission denied'}), 403

    member = app_pkg.db.query(
        """SELECT EmployeeName AS name,
                  Department  AS department,
                  EmployeeID  AS staff_id,
                  Position    AS position,
                  Level       AS level
           FROM Employee WHERE EmployeeID=?""",
        (member_id,),
        fetchone=True,
    )
    if not member:
        return jsonify({'error': 'member not found'}), 404

    excel_bytes = generate_excel_bytes(member_id, member, year, months)
    safe_name = _safe_export_name(member['name'])
    return send_file(
        io.BytesIO(excel_bytes),
        as_attachment=True,
        download_name=f'{safe_name}_Worklog_{year}.xlsx',
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )


@export_bp.route('/api/export/excel/bulk', methods=['GET'])
@elevated_required
def export_excel_bulk():
    member_ids_str = request.args.get('member_ids', '')
    year = request.args.get('year', type=int, default=date.today().year)
    months_str = request.args.get('months', '')
    months = [int(month) for month in months_str.split(',') if month.strip().isdigit()] or None

    if not member_ids_str:
        return jsonify({'error': 'member_ids required'}), 400

    member_ids = [m.strip() for m in member_ids_str.split(',') if m.strip()]

    if not member_ids:
        return jsonify({'error': 'No valid member_ids provided'}), 400

    zip_buffer = io.BytesIO()
    with zipfile.ZipFile(zip_buffer, 'w', zipfile.ZIP_DEFLATED) as archive:
        for member_id in member_ids:
            member = app_pkg.db.query(
                """SELECT EmployeeName AS name,
                          Department  AS department,
                          EmployeeID  AS staff_id,
                          Position    AS position,
                          Level       AS level
                   FROM Employee WHERE EmployeeID=?""",
                (member_id,),
                fetchone=True,
            )
            if not member:
                continue

            excel_bytes = generate_excel_bytes(member_id, member, year, months)
            safe_name = _safe_export_name(member['name'])
            archive.writestr(f'{safe_name}_Worklog_{year}.xlsx', excel_bytes)

    zip_buffer.seek(0)
    return send_file(
        zip_buffer,
        as_attachment=True,
        download_name=f'Team_Worklog_{year}.zip',
        mimetype='application/zip',
    )
